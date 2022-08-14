from selenium import webdriver
import chromedriver_binary
from selenium.webdriver.common.by import By
import time
import openpyxl

results = []

book = openpyxl.Workbook()

ws = book.active

driver = webdriver.Chrome()
driver.get("https://is.makonari.com/signin")

driver.find_element(By.NAME,'email').send_keys("ackeytoday@gmail.com")

driver.find_element(By.NAME,'password').send_keys("ぱすわーど")

driver.find_element(By.XPATH, f"/html/body/div[1]/div/div/div/form/button").click()

time.sleep(15)
for i in range(17,50):
    post_url = f"https://is.makonari.com/posts/{i}"
    driver.get(post_url)
    time.sleep(5)

    try:
        element = driver.find_element(By.CLASS_NAME, "PostDetailTitle")
        print(element.text)
        results.append([i,element.text,post_url])
    except:
        pass

for i in range(len(results)):
    ws.cell(i+1,1,value = results[i][0])
    ws.cell(i+1,2,value = results[i][1])
    ws.cell(i+1,3,value = results[i][2])

book.save('isTitles.xlsx')
